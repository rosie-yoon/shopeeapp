"""
template_analyzer.py
쇼피 템플릿 xlsx를 분석해서 auto_rules.json을 생성하는 모듈
"""

import io
import json
import re
import zipfile
import openpyxl
from openpyxl.utils import get_column_letter
from pathlib import Path
from gdrive_manager import get_gdrive_manager

CONFIG_DIR = Path(__file__).parent / "config"
CONFIG_DIR.mkdir(exist_ok=True)

AUTO_RULES_PATH = CONFIG_DIR / "auto_rules.json"
GLOBAL_RULES_PATH = CONFIG_DIR / "global_rules.json"


def _to_bytes_io(src) -> io.BytesIO:
    """
    다양한 입력 타입을 BytesIO로 변환
    src: 파일 경로(str/Path), bytes, bytearray, BytesIO, 또는 Streamlit UploadedFile
    """
    if isinstance(src, io.BytesIO):
        src.seek(0)
        return src
    if isinstance(src, (bytes, bytearray)):
        return io.BytesIO(src)
    if isinstance(src, (str, Path)):
        return io.BytesIO(Path(src).read_bytes())
    # Streamlit UploadedFile 등 read() 메서드를 가진 객체
    if hasattr(src, "read"):
        data = src.read()
        if hasattr(src, "seek"):
            src.seek(0)
        return io.BytesIO(data if isinstance(data, bytes) else data.encode())
    raise TypeError(f"지원하지 않는 입력 타입: {type(src)}")


def _fix_xlsx_bytes(src) -> io.BytesIO:
    """
    openpyxl이 읽지 못하는 확장 XML 제거 후 BytesIO 반환.
    ※ sheetViews 블록 전체를 교체하지 않고 문제 요소(<pane>, <extLst>)만 제거.
       sheetViews 전체 교체 시 tabSelected 등 속성이 사라져 Excel 복구 오류 발생.
    src: 파일 경로(str/Path), bytes, bytearray, BytesIO, Streamlit UploadedFile
    """
    zin_src = _to_bytes_io(src)

    buf = io.BytesIO()
    with zipfile.ZipFile(zin_src, "r") as zin:
        with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zout:
            for item in zin.infolist():
                data = zin.read(item.filename)
                if (item.filename.startswith("xl/worksheets/")
                        and item.filename.endswith(".xml")):
                    text = data.decode("utf-8")
                    # 1) extLst 확장 XML 제거
                    text = re.sub(r"<extLst>.*?</extLst>", "", text, flags=re.DOTALL)
                    # 2) <pane .../> 만 제거 (sheetViews 전체 X)
                    #    activePane 속성값 오류를 일으키는 원인만 정밀 제거
                    text = re.sub(r"<pane\b[^/]*/>", "", text)
                    text = re.sub(r"<pane\b[^>]*>.*?</pane>", "", text, flags=re.DOTALL)
                    # 3) <selection pane=...> 제거
                    text = re.sub(r'<selection[^>]+pane="[^"]+"[^/]*/>', "", text)
                    # 4) sheetProtection 제거 (시트 보호 해제)
                    text = re.sub(r"<sheetProtection[^/]*/>", "", text)
                    data = text.encode("utf-8")
                zout.writestr(item, data)
    buf.seek(0)
    return buf


def _load_workbook(src):
    """extLst 제거 후 openpyxl로 열기. src는 어떤 타입이든 OK."""
    buf = _fix_xlsx_bytes(src)
    return openpyxl.load_workbook(buf, data_only=True)


def _safe_name(text: str) -> str:
    """특수문자 제거, 공백→_ 변환"""
    s = re.sub(r"[^\w\s]", "", text)
    return re.sub(r"\s+", "_", s.strip())


def extract_template_info(src) -> tuple:
    """
    템플릿에서 식별 코드(C2), 대카테고리, 중카테고리를 추출

    반환: (template_code, top_category, mid_category, safe_filename)
    예시: ("100660", "Beauty", "Men's Care", "Beauty_Mens_Care.xlsx")

    src: 파일 경로(str/Path), bytes, BytesIO, 또는 Streamlit UploadedFile
    """
    wb = _load_workbook(src)
    try:
        # ── C2: 템플릿 고유 식별 코드 ──
        ws_tmpl = wb["Template"]
        template_code = str(ws_tmpl["C2"].value).strip() if ws_tmpl["C2"].value else ""
        if not template_code:
            raise ValueError("Template 시트 C2에서 식별 코드를 찾을 수 없습니다.")

        # ── Pre-order DTS Range: 대카/중카테고리 추출 ──
        # 경로 형태: "100875-Beauty/Men's Care/Bath & Body Care"
        ws = wb["Pre-order DTS Range"]
        top_cats = set()
        mid_cats = set()
        for row in ws.iter_rows(min_row=7, values_only=True):
            cat_str = row[0]
            if not cat_str:
                continue
            parts = str(cat_str).split("/")
            if len(parts) >= 2:
                # 대카: "100875-Beauty" → "Beauty" (앞의 숫자- 제거)
                top_raw = parts[0].strip()
                top_cat = re.sub(r"^\d+-", "", top_raw).strip()
                top_cats.add(top_cat)
                mid_cats.add(parts[1].strip())

        if not mid_cats:
            raise ValueError("Pre-order DTS Range 시트에서 카테고리 정보를 찾을 수 없습니다.")

        top_category = sorted(top_cats)[0]   # e.g. "Beauty"
        mid_category = sorted(mid_cats)[0]   # e.g. "Men's Care"

        # 파일명: Beauty_Mens_Care.xlsx
        filename = f"{_safe_name(top_category)}_{_safe_name(mid_category)}.xlsx"

        return template_code, top_category, mid_category, filename

    finally:
        wb.close()


def analyze_template(src) -> tuple:
    """
    템플릿 xlsx 파일을 분석해서 소카테고리별 MANDATORY 속성과 자동입력값을 반환

    src: 파일 경로(str/Path), bytes, BytesIO, 또는 Streamlit UploadedFile
    반환: (result_dict, template_code, mid_category, template_file)
    """
    # ── bytes로 한 번 읽어 두기 (여러 번 파싱해야 하므로) ──
    bio = _to_bytes_io(src)
    raw_bytes = bio.read()

    # 대/중카테고리 기반 파일명/코드 추출
    template_code, top_category, mid_category, template_file = extract_template_info(raw_bytes)

    wb = _load_workbook(raw_bytes)
    try:
        ws_template = wb["Template"]
        ws_hidden = wb["HiddenCatProps"]
        ws_attr = wb["Attribute value mapping"]

        # Step 1: Template 행1 → { col_letter: internal_key }
        # Shopee 신버전 템플릿은 키 뒤에 '|숫자|숫자' suffix가 붙음 → 제거 후 저장
        col_to_key = {}
        key_to_col = {}
        for cell in ws_template[1]:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                raw_key = str(cell.value)
                base_key = raw_key.split('|')[0]   # suffix 제거
                col_to_key[col_letter] = base_key
                key_to_col[base_key] = col_letter

        # Step 2: Template 행3 → { internal_key: display_name }
        key_to_display = {}
        for cell in ws_template[3]:
            if cell.value:
                col_letter = get_column_letter(cell.column)
                internal_key = col_to_key.get(col_letter)
                if internal_key:
                    key_to_display[internal_key] = cell.value

        # Step 3: Attribute value mapping 파싱
        attr_values = {}
        current_cat = None
        max_col = ws_attr.max_column or 1
        max_row = ws_attr.max_row or 1
        for col in range(1, max_col + 1):
            cat_val = ws_attr.cell(row=3, column=col).value
            if cat_val and cat_val not in ("Category",):
                current_cat = cat_val.strip()
            key_val = ws_attr.cell(row=1, column=col).value
            if not key_val or key_val == "et_title_global_attribute_value_mapping_attribute":
                continue
            values = []
            for row in range(7, max_row + 1):
                v = ws_attr.cell(row=row, column=col).value
                if v is not None and str(v).strip():
                    values.append(str(v).strip())
            if values and current_cat:
                attr_values[(key_val, current_cat)] = values

        # Step 4: HiddenCatProps → 소카테고리별 MANDATORY 컬럼 추출
        result = {}
        for row in ws_hidden.iter_rows(min_row=7, values_only=False):
            cat_cell = row[0].value
            if not cat_cell:
                continue
            cat_str = str(cat_cell).strip()
            if "-" not in cat_str:
                continue
            cat_id = cat_str.split("-")[0].strip()
            cat_path = "-".join(cat_str.split("-")[1:]).strip()

            mandatory_attrs = {}
            for col_idx in range(1, len(row)):
                cell_val = row[col_idx].value
                if cell_val != "MANDATORY":
                    continue
                col_letter = get_column_letter(col_idx + 1)
                internal_key = col_to_key.get(col_letter)
                if not internal_key:
                    continue
                display_name = key_to_display.get(internal_key, col_letter)
                values = attr_values.get((internal_key, cat_path), [])
                if not values:
                    for (k, _c), v in attr_values.items():
                        if k == internal_key:
                            values = v
                            break
                if values:
                    auto_value = "Others" if "Others" in values else values[-1]
                else:
                    auto_value = ""
                mandatory_attrs[internal_key] = {
                    "display": display_name,
                    "col_letter": col_letter,
                    "values": values,
                    "auto_value": auto_value,
                }

            result[cat_id] = {
                "category_path": cat_path,
                "top_category": top_category,
                "mid_category": mid_category,
                "template_code": template_code,
                "template_file": template_file,
                "mandatory_attrs": mandatory_attrs,
            }

        return result, template_code, top_category, mid_category, template_file

    finally:
        wb.close()


def save_auto_rules(new_rules: dict, template_code: str, mid_category: str):
    """
    새 분석 결과를 auto_rules.json에 저장
    - 동일 template_code(C2)의 기존 소카테고리 규칙은 모두 삭제 후 교체
    """
    existing = {}
    if AUTO_RULES_PATH.exists():
        with open(AUTO_RULES_PATH, "r", encoding="utf-8") as f:
            existing = json.load(f)

    # 동일 template_code의 기존 항목 제거
    existing = {
        cat_id: info
        for cat_id, info in existing.items()
        if info.get("template_code") != template_code
    }

    # 새 규칙 추가
    existing.update(new_rules)

    with open(AUTO_RULES_PATH, "w", encoding="utf-8") as f:
        json.dump(existing, f, ensure_ascii=False, indent=2)

    return existing


def load_auto_rules() -> dict:
    """auto_rules 로드 (구글 드라이브 우선, 로컬 fallback)"""
    try:
        gdrive = get_gdrive_manager()
        rules = gdrive.load_config_json("auto_rules.json")
        if rules:
            return rules
    except Exception as e:
        st.warning(f"구글 드라이브에서 auto_rules 로드 실패, 로컬 사용: {e}")

    # fallback: 로컬 파일
    if AUTO_RULES_PATH.exists():
        with open(AUTO_RULES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_auto_rules(new_rules: dict, template_code: str, mid_category: str):
    """auto_rules 저장 (구글 드라이브 + 로컬 백업)"""
    # 기존 규칙 로드
    existing = load_auto_rules()

    # 동일 template_code의 기존 항목 제거
    existing = {
        cat_id: info
        for cat_id, info in existing.items()
        if info.get("template_code") != template_code
    }

    # 새 규칙 추가
    existing.update(new_rules)

    # 1) 구글 드라이브에 저장
    try:
        gdrive = get_gdrive_manager()
        gdrive.save_config_json("auto_rules.json", existing)
    except Exception as e:
        st.warning(f"구글 드라이브 저장 실패: {e}")

    # 2) 로컬 백업 (fallback)
    try:
        with open(AUTO_RULES_PATH, "w", encoding="utf-8") as f:
            json.dump(existing, f, ensure_ascii=False, indent=2)
    except Exception:
        pass  # 로컬 백업 실패는 무시

    return existing


def load_global_rules() -> dict:
    """global_rules 로드 (구글 드라이브 우선, 로컬 fallback)"""
    try:
        gdrive = get_gdrive_manager()
        rules = gdrive.load_config_json("global_rules.json")
        if rules:
            return rules
    except Exception as e:
        st.warning(f"구글 드라이브에서 global_rules 로드 실패, 로컬 사용: {e}")

    # fallback: 로컬 파일
    if GLOBAL_RULES_PATH.exists():
        with open(GLOBAL_RULES_PATH, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def save_global_rules(rules: dict):
    """global_rules 저장 (구글 드라이브 + 로컬 백업)"""
    # 1) 구글 드라이브에 저장
    try:
        gdrive = get_gdrive_manager()
        gdrive.save_config_json("global_rules.json", rules)
    except Exception as e:
        st.warning(f"구글 드라이브 저장 실패: {e}")

    # 2) 로컬 백업 (fallback)
    try:
        with open(GLOBAL_RULES_PATH, "w", encoding="utf-8") as f:
            json.dump(rules, f, ensure_ascii=False, indent=2)
    except Exception:
        pass  # 로컬 백업 실패는 무시


def get_categories_by_mid(auto_rules: dict) -> dict:
    """중카테고리별로 소카테고리 그룹핑"""
    grouped = {}
    for cat_id, info in auto_rules.items():
        mid = info.get("mid_category", "Unknown")
        if mid not in grouped:
            grouped[mid] = []
        grouped[mid].append({
            "cat_id": cat_id,
            "cat_path": info["category_path"],
            "template_file": info["template_file"],
        })
    return grouped
