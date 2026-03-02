"""
file_builder.py
구글 시트 데이터 + 규칙들을 쇼피 템플릿에 이식해서 xlsx 파일을 생성하는 모듈

핵심 방식: ZIP 직접 조작 (sharedStrings 방식)
- openpyxl 저장 시 inlineStr 포맷으로 쓰면 Excel 복구 시 값이 사라지는 버그 존재
- 원본 xlsx의 sharedStrings.xml에 새 문자열을 직접 추가하고
  sheet2.xml에 t="s" 형식 행을 삽입해서 Excel 완전 호환 파일 생성
"""

import io
import re
import zipfile
import xml.etree.ElementTree as ET
import pandas as pd
from pathlib import Path
from sheet_reader import parse_category
from template_analyzer import _fix_xlsx_bytes

TEMPLATES_DIR = Path(__file__).parent / "templates"

# 구글 시트 컬럼명 → 템플릿 internal key 매핑
GSHEET_COL_TO_INTERNAL = {
    "Category":                  "ps_tmpl_mt_upload_title_category",
    "Product Name":              "ps_tmpl_mt_upload_title_product_name",
    "Product Description":       "ps_tmpl_mt_upload_title_product_description",
    "Parent SKU":                "ps_tmpl_mt_upload_title_parent_sku",
    "Variation Integration No.": "ps_tmpl_mt_upload_title_variation_integration_no",
    "Variation Name1":           "ps_tmpl_mt_upload_title_variation_1_name",
    "Option for Variation 1":    "ps_tmpl_mt_upload_title_variation_1_option",
    "Image per Variation":       "ps_tmpl_mt_upload_title_variation_1_image",
    "Global SKU Price":          "ps_tmpl_mt_upload_title_price",
    "Stock":                     "ps_tmpl_mt_upload_title_stock",
    "SKU":                       "ps_tmpl_mt_upload_title_sku",
    "Cover image":               "ps_tmpl_mt_upload_title_cover_image",
    "Item Image 1":              "ps_item_image_1",
    "Item Image 2":              "ps_item_image_2",
    "Item Image 3":              "ps_item_image_3",
    "Item Image 4":              "ps_item_image_4",
    "Item Image 5":              "ps_item_image_5",
    "Item Image 6":              "ps_item_image_6",
    "Item Image 7":              "ps_item_image_7",
    "Item Image 8":              "ps_item_image_8",
    "Weight":                    "ps_tmpl_mt_upload_title_weight",
    "Days to ship":              "ps_tmpl_mt_upload_title_dts",
    "Brand":                     "ps_tmpl_mt_upload_title_brand",
}

# 숫자로 저장해야 하는 internal key 목록
NUMERIC_KEYS = {
    "ps_tmpl_mt_upload_title_price",
    "ps_tmpl_mt_upload_title_stock",
    "ps_tmpl_mt_upload_title_weight",
    "ps_tmpl_mt_upload_title_length",
    "ps_tmpl_mt_upload_title_width",
    "ps_tmpl_mt_upload_title_height",
    "ps_tmpl_mt_upload_title_dts",
}

NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"


# ──────────────────────────────────────────────────────────────────────────────
# 내부 유틸
# ──────────────────────────────────────────────────────────────────────────────

def _col_letter(col_idx: int) -> str:
    """1-based 열 인덱스 → 컬럼 문자 (A, B, ..., Z, AA, ...)"""
    result = ""
    while col_idx > 0:
        col_idx, rem = divmod(col_idx - 1, 26)
        result = chr(65 + rem) + result
    return result


def _col_index(col_letter: str) -> int:
    """컬럼 문자 → 1-based 열 인덱스"""
    result = 0
    for ch in col_letter.upper():
        result = result * 26 + (ord(ch) - 64)
    return result


def _parse_shared_strings(ss_xml: str) -> list:
    """sharedStrings.xml 파싱 → 문자열 리스트 반환"""
    root = ET.fromstring(ss_xml)
    strings = []
    for si in root.findall(f"{{{NS}}}si"):
        # <t> 직접 자식
        t_elem = si.find(f"{{{NS}}}t")
        if t_elem is not None and t_elem.text is not None:
            strings.append(t_elem.text)
            continue
        # <r><t> rich text
        parts = []
        for r_elem in si.findall(f"{{{NS}}}r"):
            t2 = r_elem.find(f"{{{NS}}}t")
            if t2 is not None and t2.text is not None:
                parts.append(t2.text)
        strings.append("".join(parts))
    return strings


def _build_shared_strings_xml(strings: list) -> str:
    """문자열 리스트 → sharedStrings.xml 문자열"""
    count = len(strings)
    parts = [
        '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\r\n',
        f'<sst xmlns="{NS}" count="{count}" uniqueCount="{count}">',
    ]
    for s in strings:
        # XML 특수문자 이스케이프
        escaped = (str(s)
                   .replace("&", "&amp;")
                   .replace("<", "&lt;")
                   .replace(">", "&gt;")
                   .replace('"', "&quot;"))
        # 앞뒤 공백 보존 속성
        space_attr = ' xml:space="preserve"' if escaped != escaped.strip() else ""
        parts.append(f"<si><t{space_attr}>{escaped}</t></si>")
    parts.append("</sst>")
    return "".join(parts)


def _get_key_to_col(sheet2_xml: str) -> dict:
    """sheet2.xml row1에서 { internal_key: col_letter } 매핑"""
    rows1 = re.findall(r'<row[^>]*\br="1"[^>]*>.*?</row>', sheet2_xml, re.DOTALL)
    if not rows1:
        return {}
    mapping = {}
    for m in re.finditer(r'<c\s+r="([A-Z]+)1"[^>]*>\s*(?:<v>(\d+)</v>|<is><t[^>]*>([^<]*)</t></is>)', rows1[0]):
        col_letter = m.group(1)
        # sharedString 인덱스 or inlineStr
        mapping[col_letter] = m.group(2) or m.group(3)  # 임시: 인덱스 문자열 저장
    return mapping


def _strip_key_suffix(raw_key: str) -> str:
    """키 값에 붙은 |숫자|숫자 suffix 제거. 예: 'ps_tmpl_mt_upload_title_category|1|0' → 'ps_tmpl_mt_upload_title_category'"""
    return raw_key.split('|')[0] if '|' in raw_key else raw_key


def _get_key_to_col_from_shared(sheet2_xml: str, shared: list) -> dict:
    """row1 셀 값(sharedString 또는 inlineStr)으로 { internal_key: col_letter } 매핑
    Shopee 템플릿 버전에 따라 키 뒤에 '|1|0' 같은 suffix가 붙는 경우도 처리함.
    """
    rows1 = re.findall(r'<row[^>]*\br="1"[^>]*>.*?</row>', sheet2_xml, re.DOTALL)
    if not rows1:
        return {}
    mapping = {}
    # sharedString 참조: <c r="A1" ... t="s"><v>N</v></c>
    for m in re.finditer(r'<c\s+r="([A-Z]+)1"[^>]*t="s"[^>]*>\s*<v>(\d+)</v>', rows1[0]):
        col_letter = m.group(1)
        idx = int(m.group(2))
        if idx < len(shared):
            raw_key = shared[idx]
            base_key = _strip_key_suffix(raw_key)
            mapping[base_key] = col_letter          # suffix 없는 키로 저장
            if raw_key != base_key:
                mapping[raw_key] = col_letter       # suffix 있는 원본도 저장
    # inlineStr: <c r="A1" ... t="inlineStr"><is><t>...</t></is></c>
    for m in re.finditer(r'<c\s+r="([A-Z]+)1"[^>]*t="inlineStr"[^>]*>\s*<is><t[^>]*>([^<]+)</t></is>', rows1[0]):
        col_letter = m.group(1)
        raw_key = m.group(2)
        base_key = _strip_key_suffix(raw_key)
        mapping[base_key] = col_letter
        if raw_key != base_key:
            mapping[raw_key] = col_letter
    return mapping


def _get_status_map_from_hidden(hidden_xml: str, shared: list, cat_id: str) -> dict:
    """
    HiddenCatProps (sheet7) XML에서 cat_id 행의 { col_letter: status } 반환
    status: MANDATORY / OPTIONAL / IRRELEVANT / None
    """
    target_prefix = f"{cat_id}-"
    # row 7 이후 모든 row 순회
    for row_m in re.finditer(r'<row\b[^>]*>(.*?)</row>', hidden_xml, re.DOTALL):
        row_xml = row_m.group(0)
        row_content = row_m.group(1)
        # 첫 번째 셀 A열 값 확인
        first_cell = re.search(
            r'<c\s+r="A\d+"[^>]*t="s"[^>]*>\s*<v>(\d+)</v>', row_content
        )
        if not first_cell:
            continue
        val_idx = int(first_cell.group(1))
        if val_idx >= len(shared):
            continue
        cell_val = shared[val_idx]
        if not str(cell_val).startswith(target_prefix):
            continue
        # 해당 행 모든 셀 상태 수집
        status_by_col = {}
        for cm in re.finditer(
            r'<c\s+r="([A-Z]+)\d+"[^>]*(?:t="s")?[^>]*>\s*(?:<v>(\d+)</v>)?', row_content
        ):
            col_letter = cm.group(1)
            v = cm.group(2)
            if v is not None:
                sidx = int(v)
                status_by_col[col_letter] = shared[sidx] if sidx < len(shared) else None
            else:
                status_by_col[col_letter] = None
        return status_by_col
    return {}


def _make_cell_xml(col_letter: str, row_num: int, value, shared: list, style: str = "0") -> str:
    """
    셀 XML 생성
    - 숫자: t 속성 없음, <v>숫자</v>
    - 문자열: t="s", sharedStrings에 추가 후 인덱스 참조
    """
    cell_ref = f"{col_letter}{row_num}"

    # 숫자 처리
    try:
        if isinstance(value, (int, float)):
            num_val = float(value)
        else:
            num_val = float(str(value).replace(",", "").strip())
        # 정수인 float은 int로 변환: 100.0→100, 3.0→3
        # Shopee ParseUint 오류 방지 (Stock, DTS, Brand ID 등)
        if num_val == int(num_val):
            return f'<c r="{cell_ref}" s="{style}"><v>{int(num_val)}</v></c>'
        else:
            return f'<c r="{cell_ref}" s="{style}"><v>{num_val}</v></c>'
    except (ValueError, TypeError):
        pass

    # 문자열 처리 → sharedStrings 방식
    str_val = str(value)
    # 기존 shared에 있는지 확인
    if str_val in shared:
        idx = shared.index(str_val)
    else:
        idx = len(shared)
        shared.append(str_val)
    return f'<c r="{cell_ref}" s="{style}" t="s"><v>{idx}</v></c>'


def _build_row_xml(row_num: int, cells_xml: list) -> str:
    """행 XML 생성"""
    inner = "".join(cells_xml)
    return f'<row r="{row_num}" spans="1:112">{inner}</row>'


# ──────────────────────────────────────────────────────────────────────────────
# 공개 API (기존과 동일한 시그니처 유지)
# ──────────────────────────────────────────────────────────────────────────────

def get_template_col_map(ws_template) -> dict:
    """(하위 호환) openpyxl worksheet → { internal_key: col_letter }"""
    from openpyxl.utils import get_column_letter
    key_to_col = {}
    for cell in ws_template[1]:
        if cell.value:
            key_to_col[cell.value] = get_column_letter(cell.column)
    return key_to_col


def get_template_status_map(ws_hidden, key_to_col: dict, cat_id: str, cat_path: str) -> dict:
    """(하위 호환) openpyxl worksheet 기반 상태맵 반환"""
    from openpyxl.utils import get_column_letter
    target_prefix = f"{cat_id}-"
    for row in ws_hidden.iter_rows(min_row=7, max_row=ws_hidden.max_row, values_only=False):
        cell_val = row[0].value
        if not cell_val:
            continue
        if str(cell_val).startswith(target_prefix):
            status_by_col = {}
            for col_idx in range(1, len(row)):
                col_letter = get_column_letter(col_idx + 1)
                status_by_col[col_letter] = row[col_idx].value
            return status_by_col
    return {}


def build_file(
    template_path: str,
    group_df: pd.DataFrame,
    cat_id: str,
    auto_rules: dict,
    global_rules: dict,
) -> bytes:
    """
    ZIP 직접 조작 방식으로 sharedString 포맷 유지하며 데이터 삽입
    Excel 열었을 때 값이 사라지는 inlineStr 문제 완전 해결
    """
    # ── 1. 원본 바이트 읽기 ──
    with open(template_path, "rb") as f:
        raw = f.read()

    # ── 2. ZIP 파싱 + 전체 시트 XML 정리 ──
    # activePane="bottom_left" 는 OOXML 스펙 위반 → "bottomLeft" 로 수정
    # (원본 Shopee 템플릿 버그 - 이것이 Excel 복구 오류 sheet2/4/5 원인)
    with zipfile.ZipFile(io.BytesIO(raw)) as zf:
        names = zf.namelist()
        file_map = {}
        for n in names:
            data = zf.read(n)
            if n.startswith("xl/worksheets/") and n.endswith(".xml"):
                text = data.decode("utf-8")
                # activePane 값 수정 (bottom_left → bottomLeft 등)
                text = text.replace('activePane="bottom_left"', 'activePane="bottomLeft"')
                text = text.replace('activePane="top_left"', 'activePane="topLeft"')
                text = text.replace('activePane="bottom_right"', 'activePane="bottomRight"')
                text = text.replace('activePane="top_right"', 'activePane="topRight"')
                data = text.encode("utf-8")
            file_map[n] = data

    ss_xml    = file_map["xl/sharedStrings.xml"].decode("utf-8")
    sheet2_xml = file_map["xl/worksheets/sheet2.xml"].decode("utf-8")

    # sheet2.xml 에만 추가 수정: sheetProtection + extLst 제거
    # sheetProtection 은 self-closing 또는 non-self-closing 모두 처리
    sheet2_xml = re.sub(r"<sheetProtection\b[^>]*/>", "", sheet2_xml)
    sheet2_xml = re.sub(r"<sheetProtection\b[^>]*>.*?</sheetProtection>", "", sheet2_xml, flags=re.DOTALL)
    sheet2_xml = re.sub(r"<extLst>.*?</extLst>", "", sheet2_xml, flags=re.DOTALL)

    # HiddenCatProps 시트 파일 찾기 (sheet7.xml)
    hidden_xml = file_map.get("xl/worksheets/sheet7.xml", b"").decode("utf-8")

    # ── 4. sharedStrings 파싱 ──
    shared = _parse_shared_strings(ss_xml)

    # ── 5. internal_key → col_letter 매핑 (row1) ──
    key_to_col = _get_key_to_col_from_shared(sheet2_xml, shared)

    # ── 6. cat_id 상태맵 ──
    status_map = _get_status_map_from_hidden(hidden_xml, shared, cat_id)
    cat_info   = auto_rules.get(cat_id, {})

    DATA_START_ROW = 7

    # ── 7. 삽입할 행 XML 생성 ──
    new_rows_xml = []

    for row_idx, (_, sheet_row) in enumerate(group_df.iterrows()):
        write_row = DATA_START_ROW + row_idx
        cells_xml = []

        # 데이터 셀 딕셔너리: { col_letter: value }
        cell_data = {}

        # ① 구글시트 데이터
        for gs_col, internal_key in GSHEET_COL_TO_INTERNAL.items():
            value = sheet_row.get(gs_col, "")
            if value == "" or (isinstance(value, float) and pd.isna(value)):
                continue
            col_letter = key_to_col.get(internal_key)
            if col_letter:
                cell_data[col_letter] = value

        # ② auto_rules mandatory 자동값
        for internal_key, attr_info in cat_info.get("mandatory_attrs", {}).items():
            auto_val = attr_info.get("auto_value", "")
            if auto_val:
                col_letter = key_to_col.get(internal_key)
                if col_letter:
                    cell_data[col_letter] = auto_val

        # ③ global_rules 적용
        for internal_key, rule in global_rules.items():
            value = rule.get("value", "")
            apply_when = rule.get("apply_when", "exists")
            if not value:
                continue
            col_letter = key_to_col.get(internal_key)
            if not col_letter:
                continue
            col_status = status_map.get(col_letter)
            if apply_when == "exists" and col_status != "IRRELEVANT":
                cell_data[col_letter] = value
            elif apply_when == "mandatory" and col_status == "MANDATORY":
                cell_data[col_letter] = value

        # 열 순서 정렬 후 셀 XML 생성
        sorted_cols = sorted(cell_data.keys(), key=lambda c: _col_index(c))
        for col_letter in sorted_cols:
            value = cell_data[col_letter]
            # 숫자 키 여부로 스타일 결정 (스타일 0 기본)
            internal_key = None
            for k, v in key_to_col.items():
                if v == col_letter:
                    internal_key = k
                    break
            cell_xml = _make_cell_xml(col_letter, write_row, value, shared, style="0")
            cells_xml.append(cell_xml)

        if cells_xml:
            new_rows_xml.append(_build_row_xml(write_row, cells_xml))

    # ── 8. sheet2.xml에 행 삽입 ──
    # </sheetData> 바로 앞에 새 행 삽입
    rows_str = "".join(new_rows_xml)
    new_sheet2 = sheet2_xml.replace("</sheetData>", rows_str + "</sheetData>", 1)

    # ── 9. sharedStrings.xml 재생성 ──
    new_ss_xml = _build_shared_strings_xml(shared)

    # ── 10. 새 ZIP 생성 ──
    out_buf = io.BytesIO()
    with zipfile.ZipFile(out_buf, "w", zipfile.ZIP_DEFLATED) as zf_out:
        for name in names:
            if name == "xl/sharedStrings.xml":
                zf_out.writestr(name, new_ss_xml.encode("utf-8"))
            elif name == "xl/worksheets/sheet2.xml":
                zf_out.writestr(name, new_sheet2.encode("utf-8"))
            else:
                zf_out.writestr(name, file_map[name])

    out_buf.seek(0)
    return out_buf.getvalue()


def build_all_files(
    groups: dict,
    auto_rules: dict,
    global_rules: dict,
) -> tuple:
    """
    카테고리별 그룹 전체 처리 → { filename: bytes }, skipped_list 반환
    """
    results = {}
    skipped = []

    for cat_id, group_df in groups.items():
        if cat_id == "unknown":
            skipped.append("카테고리 파싱 실패 상품")
            continue

        cat_info = auto_rules.get(cat_id)
        if not cat_info:
            skipped.append(f"카테고리 {cat_id} (템플릿 없음)")
            continue

        template_file = cat_info.get("template_file")
        template_path = TEMPLATES_DIR / template_file
        if not template_path.exists():
            skipped.append(f"카테고리 {cat_id} ({template_file} 파일 없음)")
            continue

        try:
            file_bytes = build_file(
                str(template_path),
                group_df,
                cat_id,
                auto_rules,
                global_rules,
            )
            safe_name = (
                cat_info.get("category_path", cat_id)
                .replace("/", "_")
                .replace(" ", "_")
            )
            filename = f"shopee_{safe_name}_{cat_id}.xlsx"
            results[filename] = file_bytes
        except Exception as e:
            skipped.append(f"카테고리 {cat_id} 처리 오류: {e}")

    return results, skipped
